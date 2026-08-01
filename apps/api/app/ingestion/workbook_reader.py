"""Reads the two fixed ANSADE/CN national-accounts workbooks by exact
worksheet name, using the known, fixed row order (see taxonomy.py) rather
than fuzzy header matching. Each workbook has years across a header row and
one indicator per subsequent row (see docs/DATA_DICTIONARY.md)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from app.ingestion.label_clean import clean_label
from app.ingestion.normalize import normalize_text
from app.ingestion.numeric import NotNumericError, parse_numeric
from app.ingestion.taxonomy import IndicatorRowSpec

_YEAR_RE = re.compile(r"^(19|20)\d{2}$")


class WorkbookStructureError(ValueError):
    """Raised when the workbook doesn't match the expected fixed layout."""


@dataclass
class YearCell:
    year: int
    value: float | None
    original_value: str | None
    is_missing: bool
    quality_flag: str  # ok | missing | nonnumeric
    column_letter: str


@dataclass
class RowResult:
    spec: IndicatorRowSpec
    raw_label: str
    source_row: int
    label_matches_expected: bool
    cells: list[YearCell] = field(default_factory=list)


@dataclass
class WorkbookReadResult:
    worksheet_name: str
    title: str
    year_header_row: int
    start_year: int | None
    end_year: int | None
    rows: list[RowResult]
    warnings: list[str]


def _find_worksheet(wb: openpyxl.Workbook, expected_name: str):
    target = expected_name.strip()
    for name in wb.sheetnames:
        if name.strip() == target:
            return wb[name]
    raise WorkbookStructureError(
        f"Worksheet '{expected_name}' not found. Available sheets: {wb.sheetnames}"
    )


def _find_year_header_row(ws) -> tuple[int, dict[int, int]]:
    """Returns (row_index, {year: column_index}) for the first row where at
    least 10 cells parse as plausible 4-digit years."""
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        year_cols: dict[int, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            raw = ws.cell(row=row_idx, column=col_idx).value
            if raw is None:
                continue
            text = str(raw).strip()
            if _YEAR_RE.match(text):
                year_cols[int(text)] = col_idx
        if len(year_cols) >= 10:
            return row_idx, year_cols
    raise WorkbookStructureError("No year header row found in the first 10 rows.")


def _parse_cell(raw) -> tuple[float | None, str | None, bool, str]:
    if raw is None:
        return None, None, True, "missing"
    text = str(raw).replace("\xa0", " ").strip()
    if text == "" or text.upper() == "NA":
        return None, None, True, "missing"
    try:
        value = parse_numeric(text)
        return value, text, False, "ok"
    except NotNumericError:
        return None, text, False, "nonnumeric"


def read_workbook(
    file_path: str | Path,
    expected_sheet_name: str,
    row_specs: tuple[IndicatorRowSpec, ...],
) -> WorkbookReadResult:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = _find_worksheet(wb, expected_sheet_name)

    title = str(ws.cell(row=1, column=1).value or "").strip()
    header_row, year_cols = _find_year_header_row(ws)

    warnings: list[str] = []
    rows: list[RowResult] = []

    spec_idx = 0
    row_idx = header_row + 1
    while spec_idx < len(row_specs) and row_idx <= ws.max_row:
        raw_label = ws.cell(row=row_idx, column=1).value
        if raw_label is None or str(raw_label).strip() == "":
            row_idx += 1
            continue
        raw_label_str = str(raw_label)
        if normalize_text(raw_label_str).startswith("source"):
            break  # reached the "Source : ANSADE/CN" footer row

        spec = row_specs[spec_idx]
        cleaned = clean_label(raw_label_str)
        normalized = normalize_text(cleaned)
        matches = spec.label_fragment in normalized
        if not matches:
            warnings.append(
                f"Row {row_idx}: label '{raw_label_str}' (normalized '{normalized}') "
                f"did not contain expected fragment '{spec.label_fragment}' for "
                f"indicator '{spec.code}'. Row consumed positionally regardless."
            )

        cells: list[YearCell] = []
        for year in sorted(year_cols):
            col_idx = year_cols[year]
            raw_value = ws.cell(row=row_idx, column=col_idx).value
            value, original_value, is_missing, flag = _parse_cell(raw_value)
            cells.append(
                YearCell(
                    year=year,
                    value=value,
                    original_value=original_value,
                    is_missing=is_missing,
                    quality_flag=flag,
                    column_letter=openpyxl.utils.get_column_letter(col_idx),
                )
            )

        rows.append(
            RowResult(
                spec=spec,
                raw_label=raw_label_str,
                source_row=row_idx,
                label_matches_expected=matches,
                cells=cells,
            )
        )
        spec_idx += 1
        row_idx += 1

    if spec_idx < len(row_specs):
        missing_codes = [s.code for s in row_specs[spec_idx:]]
        raise WorkbookStructureError(
            f"Expected {len(row_specs)} indicator rows in '{expected_sheet_name}' but "
            f"only found {spec_idx} before the sheet ended. Missing: {missing_codes}"
        )

    years = sorted(year_cols)
    return WorkbookReadResult(
        worksheet_name=ws.title,
        title=title,
        year_header_row=header_row,
        start_year=years[0] if years else None,
        end_year=years[-1] if years else None,
        rows=rows,
        warnings=warnings,
    )
